import requests
from bs4 import BeautifulSoup
import openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Top rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank & Name ','Year of Release','IMDB Rating'])


# Define headers to mimic a real browser request
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

try:
    source = requests.get('https://www.imdb.com/chart/top/', headers=headers)
    source.raise_for_status()  # Check if the request was successful

    soup = BeautifulSoup(source.text, 'html.parser')
    
    movies = soup.find('ul',class_="ipc-metadata-list ipc-metadata-list--dividers-between sc-e22973a9-0 khSCXM compact-list-view ipc-metadata-list--base").find_all('li')
    
    for movie in movies:

        name=movie.find('div',class_='ipc-title ipc-title--base ipc-title--title ipc-title-link-no-icon ipc-title--on-textPrimary sc-3713cfda-2 fSzZES cli-title with-margin').a.text
        #rank=movie.find('div',class_='ipc-title ipc-title--base ipc-title--title ipc-title-link-no-icon ipc-title--on-textPrimary sc-3713cfda-2 fSzZES cli-title with-margin').a.get_text(strip=True).split('.')[0]

        year=movie.find('div',class_='sc-d5ea4b9d-6 hBxwRe cli-title-metadata').span.text

        rating=movie.find('span',class_="ipc-rating-star--rating").text
        print(name,year,rating)
        sheet.append([name,year,rating])

except Exception as e:
    print(f"Error: {e}")
    
excel.save('IMDB Movie Rating.xlsx')